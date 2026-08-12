from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font
import tkinter as tk
from tkinter import ttk, messagebox
import pyperclip
import os
import sys
import shutil


def is_frozen():
    return getattr(sys, "frozen", False)


def bundle_dir():
    """Read-only resources bundled by PyInstaller (or project root in dev)."""
    if is_frozen():
        return sys._MEIPASS
    return os.path.abspath(os.path.dirname(__file__))


def app_dir():
    """Writable directory next to the .exe (or project root in dev)."""
    if is_frozen():
        return os.path.dirname(sys.executable)
    return os.path.abspath(os.path.dirname(__file__))


def resource_path(relative_path):
    """Absolute path to a bundled resource (themes, icon, template)."""
    return os.path.join(bundle_dir(), relative_path)


def create_empty_workbook(path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Passwords"
    for col, header in enumerate(["Name", "Password", "Email", "Remarks"], start=1):
        cell = ws.cell(1, col, header)
        cell.font = Font(bold=True)
    ws2 = wb.create_sheet("Shortcuts")
    ws2["A1"] = ""
    ws2["A2"] = ""
    ws2["A3"] = ""
    wb.save(path)


def ensure_data_file():
    """
    Keep passwords in a writable Excel file NEXT TO the app, never inside _internal.
    Never overwrite an existing data2.xlsx (user passwords stay local).
    """
    data_dir = os.path.join(app_dir(), "assets")
    os.makedirs(data_dir, exist_ok=True)
    data_file = os.path.join(data_dir, "data2.xlsx")

    if os.path.exists(data_file):
        return data_file

    template = resource_path(os.path.join("assets", "data2.template.xlsx"))
    if os.path.exists(template):
        shutil.copy2(template, data_file)
    else:
        create_empty_workbook(data_file)

    return data_file


def printallsheet():
    for row in ws.iter_rows(min_row=1, min_col=1):
        for cell in row:
            print(cell.value, end=" ")
        print()


def printlist(list):
    for i in list:
        print(i)


def OneRow(number):
    myrow = []
    for row in ws.iter_rows(min_row=number, min_col=1, max_row=number):
        for cell in row:
            myrow.append(cell.value)
    return myrow


def resetSheet():
    ask = messagebox.askquestion("Remove All", "Are you sure?")
    if ask == "yes":
        ws.delete_rows(1, ws.max_row)
        ws["A1"].value = "Name"
        ws["A1"].font = Font(bold=True)
        ws["B1"].value = "Password"
        ws["B1"].font = Font(bold=True)
        ws["C1"].value = "Email"
        ws["C1"].font = Font(bold=True)
        ws["D1"].value = "Remarks"
        ws["D1"].font = Font(bold=True)
        printallsheet()
        updatesheetname()
        cmb4.set("")
        resetInfoEntry()
    else:
        return


def names():
    lnames = []
    for row in ws:
        if row[0].value == "Name":
            continue
        if row[0].value is None:
            continue
        lnames.append(row[0].value)
    return lnames


def insertdata():
    name = name_entry.get()
    passw = pass_entry.get()
    email = email_entry.get()
    remark = re_entry.get()
    ist = isthere(name)
    rowval = [name, passw, email, remark]
    if name != " " and name is not None and name != "" and passw != "" and email != "":
        if ist != True:
            ws.append(rowval)
            print(names())
            updatesheetname()
            resetentrys()
        else:
            messagebox.showerror("Error", "This Name is already exists")
    else:
        messagebox.showerror("Error", "The Name, Password and Email are required field")


def resetentrys():
    name_entry.delete("0", "end")
    pass_entry.delete("0", "end")
    email_entry.delete("0", "end")
    re_entry.delete("0", "end")


def updatesheetname():
    cmb4["values"] = names()
    xfile.save(file2)


def isthere(name):
    allnames = names()
    if name in allnames:
        return True
    lowered = [n.lower() for n in allnames if isinstance(n, str)]
    if isinstance(name, str) and name.lower() in lowered:
        return True
    return False


def showpass():
    if pass_entry.cget("show") == "*":
        pass_entry.config(show="")
    else:
        pass_entry.config(show="*")


def findrowbyname(name):
    allnames = names()
    if name in allnames:
        return allnames.index(name) + 2
    if isinstance(name, str):
        for i, n in enumerate(allnames):
            if isinstance(n, str) and n.lower() == name.lower():
                return i + 2
    return None


def search(event):
    val = event.widget.get()
    allnames = names()
    if val == "":
        cmb4["values"] = allnames
    else:
        data = []
        for item in allnames:
            if item is not None and val.lower() in str(item).lower():
                data.append(item)
        cmb4["values"] = data


def rowbynamebyname(event):
    rownum = findrowbyname(cmb4.get())
    if not rownum:
        return
    row = OneRow(rownum)
    setInfoEntry(row[0], row[1], row[2], row[3])
    cmb4.set("")


def setInfoEntry(name, passw, email, re):
    InNameStr.set(name if name is not None else "")
    InPassStr.set(passw if passw is not None else "")
    InEmailStr.set(email if email is not None else "")
    InReStr.set(re if re is not None else "")


def resetInfoEntry():
    setInfoEntry("", "", "", "")


def infoshowpass():
    if in_pass_entry.cget("show") == "*":
        in_pass_entry.config(show="")
    else:
        in_pass_entry.config(show="*")


def SwitchMode():
    if modecheck.instate(["selected"]):
        style.theme_use("forest-light")
        changeInfoBtnColerWhite()
    else:
        style.theme_use("forest-dark")
        changeInfoBtnColerBlack()


def deleteSpRow():
    ask = messagebox.askquestion("Remove All", "Are you sure?")
    if ask == "yes":
        if isthere(InNameStr.get()):
            row = findrowbyname(InNameStr.get())
            if row:
                ws.delete_rows(row, 1)
                updatesheetname()
                resetInfoEntry()
        else:
            return
    else:
        return


def changeInfoBtnColerWhite():
    in_btnCopyPass.configure(bg="#dcdcdc")
    in_btnCopyPass.configure(fg="black")
    in_btnCopyEmail.configure(bg="#dcdcdc")
    in_btnCopyEmail.configure(fg="black")
    in_btnRemoveItem.configure(bg="#dcdcdc")
    in_btnRemoveItem.configure(fg="black")


def changeInfoBtnColerBlack():
    in_btnCopyPass.configure(bg="#5c5c5c")
    in_btnCopyPass.configure(fg="white")
    in_btnCopyEmail.configure(bg="#5c5c5c")
    in_btnCopyEmail.configure(fg="white")
    in_btnRemoveItem.configure(bg="#5c5c5c")
    in_btnRemoveItem.configure(fg="white")


def getShortInfoRow():
    infoProp = []
    if ws2["A1"].value is None:
        infoProp.append("")
    else:
        infoProp.append(ws2["A1"].value)
    if ws2["A2"].value is None:
        infoProp.append("")
    else:
        infoProp.append(ws2["A2"].value)

    if ws2["A3"].value is None:
        infoProp.append("")
    else:
        infoProp.append(ws2["A3"].value)

    return infoProp


def changeShortInfo():
    def aplyy():
        ws2["A1"] = MainEntey.get()
        ws2["A2"] = SecEntry.get()
        ws2["A3"] = IdEntry.get()
        xfile.save(file2)
        toplevel.destroy()

    toplevel = tk.Toplevel(root)
    toplevel.title("Change")
    toplevel.grab_set()
    x = root.winfo_x()
    y = root.winfo_y()
    toplevel.geometry("+%d+%d" % (x + 200, y + 200))
    toplevel.geometry("235x170")
    try:
        toplevel.iconbitmap(resource_path(os.path.join("assets", "passico.ico")))
    except Exception:
        pass
    toplevel.resizable(False, False)
    toplevel.wm_transient(root)

    MainE = ttk.Label(toplevel, text="Main Email")
    SecE = ttk.Label(toplevel, text="Second Email")
    IdLeb = ttk.Label(toplevel, text="Identity Number")

    MainE.grid(row=0, column=0, sticky="w", pady=3, padx=3)
    SecE.grid(row=1, column=0, sticky="w", pady=3, padx=3)
    IdLeb.grid(row=2, column=0, sticky="w", pady=3, padx=3)

    MainEntey = ttk.Entry(toplevel, width=14)
    SecEntry = ttk.Entry(toplevel, width=14)
    IdEntry = ttk.Entry(toplevel, width=14)

    MainEntey.grid(row=0, column=1, sticky="w", pady=5, padx=3)
    SecEntry.grid(row=1, column=1, sticky="w", pady=5, padx=3)
    IdEntry.grid(row=2, column=1, sticky="w", pady=5, padx=3)
    Srow = getShortInfoProp()

    MainEntey.insert(0, Srow[0])
    SecEntry.insert(0, Srow[1])
    IdEntry.insert(0, Srow[2])

    btnCanc = ttk.Button(toplevel, text="Cancel", command=lambda: toplevel.destroy())
    btnAply = ttk.Button(toplevel, text="Apply", command=aplyy)
    btnCanc.grid(row=3, column=0, padx=10, sticky="e")
    btnAply.grid(row=3, column=1, sticky="w")
    xfile.save(file2)


file2 = ensure_data_file()
xfile = load_workbook(file2)
sheets = xfile.sheetnames
ws = xfile[sheets[0]]
ws2 = xfile[sheets[1]]
xfile.save(file2)


root = tk.Tk()
root.eval("tk::PlaceWindow . center")
root.title("PassPro")
root.geometry("650x400")
root.resizable(False, False)
style = ttk.Style(root)
try:
    root.iconbitmap(resource_path(os.path.join("assets", "passico.ico")))
except Exception:
    pass
root.tk.call("source", resource_path(os.path.join("assets", "forest-light.tcl")))
root.tk.call("source", resource_path(os.path.join("assets", "forest-dark.tcl")))
style.theme_use("forest-dark")


# Thebigframe
frame = ttk.Frame(root)
frame.pack(fill="both", expand="yes")
# search frame
search_frame = ttk.LabelFrame(frame, text="Search")
search_frame.pack(side="left", fill="y", padx=20, pady=15)

cmb4 = ttk.Combobox(search_frame, value=names(), width=35)
cmb4.set("")
cmb4.pack(padx=5)
cmb4.bind("<KeyRelease>", search)
cmb4.bind("<<ComboboxSelected>>", rowbynamebyname)

info_frame = ttk.LabelFrame(search_frame, text="info")
info_frame.pack(fill="both", expand="yes", padx=10, pady=(200, 10))
in_name_label = ttk.Label(info_frame, text="Name")
in_pass_label = ttk.Label(info_frame, text="Password")
in_email_label = ttk.Label(info_frame, text="Email")
in_re_label = ttk.Label(info_frame, text="Remarks")

in_name_label.grid(row=0, column=0, sticky="w", padx=5)
in_pass_label.grid(row=0, column=1, sticky="w", padx=5)
in_email_label.grid(row=0, column=2, sticky="w", padx=5)
in_re_label.grid(row=0, column=3, sticky="w", padx=5)

InNameStr = tk.StringVar()
InPassStr = tk.StringVar()
InEmailStr = tk.StringVar()
InReStr = tk.StringVar()

in_name_entry = ttk.Entry(
    info_frame, width=8, font=("Helvercia", 8), state="readonly", textvariable=InNameStr
)
in_pass_entry = ttk.Entry(
    info_frame,
    show="*",
    width=8,
    font=("Helvercia", 8),
    state="readonly",
    textvariable=InPassStr,
)
in_email_entry = ttk.Entry(
    info_frame, width=8, font=("Helvercia", 8), state="readonly", textvariable=InEmailStr
)
in_re_entry = ttk.Entry(
    info_frame, width=8, font=("Helvercia", 8), state="readonly", textvariable=InReStr
)
varpas = tk.BooleanVar()
in_check_showpass = ttk.Checkbutton(info_frame, command=infoshowpass, variable=varpas)
in_btnCopyPass = tk.Button(
    info_frame,
    text="copy",
    height=1,
    width=3,
    command=lambda: pyperclip.copy(in_pass_entry.get()),
    bg="#5c5c5c",
)
in_btnCopyEmail = tk.Button(
    info_frame,
    text="copy",
    height=1,
    width=5,
    command=lambda: pyperclip.copy(in_email_entry.get()),
    bg="#5c5c5c",
)
in_btnRemoveItem = tk.Button(
    info_frame, text="Remove", height=1, width=5, command=deleteSpRow, bg="#5c5c5c"
)

in_name_entry.grid(row=1, column=0, padx=(3, 1))
in_pass_entry.grid(row=1, column=1)
in_email_entry.grid(row=1, column=2)
in_re_entry.grid(row=1, column=3, padx=(2, 0))
in_btnRemoveItem.grid(row=2, column=0)
in_check_showpass.grid(row=2, column=1, sticky="e", padx=(6, 0))
in_btnCopyPass.grid(row=2, column=1, sticky="w", pady=3, padx=(3, 0))
in_btnCopyEmail.grid(row=2, column=2, padx=2)

new_frame = ttk.LabelFrame(frame, text="new")
new_frame.pack(fill="y", pady=(15, 0))

name_label = ttk.Label(new_frame, text="Name")
pass_label = ttk.Label(new_frame, text="Password")
email_label = ttk.Label(new_frame, text="Email")
re_label = ttk.Label(new_frame, text="Remarks")

name_label.grid(row=0, column=0, sticky="w", padx=5)
pass_label.grid(row=1, column=0, sticky="w", padx=5)
email_label.grid(row=2, column=0, sticky="w", padx=5)
re_label.grid(row=3, column=0, sticky="w", padx=5)

name_entry = ttk.Entry(new_frame)
pass_entry = ttk.Entry(new_frame, show="*")
email_entry = ttk.Entry(new_frame)
re_entry = ttk.Entry(new_frame)

varpas2 = tk.BooleanVar()
pass_chb = ttk.Checkbutton(new_frame, command=showpass, variable=varpas2)

name_entry.grid(row=0, column=1, pady=2)
pass_entry.grid(row=1, column=1, pady=2)
pass_chb.grid(row=1, column=2)
email_entry.grid(row=2, column=1, pady=2)
re_entry.grid(row=3, column=1, pady=2)

btnadd = ttk.Button(new_frame, text="Insert", command=insertdata)
btnreentry = ttk.Button(new_frame, text="Reset", command=resetentrys)
btnadd.grid(row=4, column=0, padx=5, pady=5, sticky="w")
btnreentry.grid(row=4, column=1, padx=5, pady=5, sticky="w")


short_frame = ttk.LabelFrame(frame, text="shortcut")
short_frame.pack(fill="y", expand="yes", pady=15)
btnresetALL = ttk.Button(short_frame, text="Remove All", command=resetSheet, width=16)
btnCopyMainEmail = ttk.Button(
    short_frame,
    text="copy main email",
    command=lambda: pyperclip.copy(str(ws2["A1"].value)),
    width=16,
)
btnCopyMyEmail = ttk.Button(
    short_frame,
    text="copy second email",
    command=lambda: pyperclip.copy(str(ws2["A2"].value)),
    width=16,
)
btnCopyId = ttk.Button(
    short_frame,
    text="copy Id",
    command=lambda: pyperclip.copy(str(ws2["A3"].value)),
    width=16,
)
modecheck = ttk.Checkbutton(short_frame, text="mode", style="Switch", command=SwitchMode)

# change
btnChange = ttk.Button(short_frame, text="Change", command=changeShortInfo, width=16)

btnresetALL.grid(row=0, column=0, padx=8, pady=8)
btnCopyMainEmail.grid(row=0, column=1, padx=(0, 8), pady=8)
btnCopyMyEmail.grid(row=1, column=0, padx=8)
btnCopyId.grid(row=1, column=1, padx=(0, 8))
modecheck.grid(row=2, column=0, pady=8, padx=10, sticky="w")
btnChange.grid(row=2, column=1, pady=(6, 0), padx=(0, 8))


root.mainloop()
